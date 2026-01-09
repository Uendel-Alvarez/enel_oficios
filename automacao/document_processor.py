# automacao/document_processor.py

import os
from docx import Document
import openpyxl
from PyPDF2 import PdfReader

# Funções auxiliares para ler/extrair texto de diferentes tipos de arquivo
def extract_text_from_pdf(file_path):
    try:
        reader = PdfReader(file_path)
        text = ""
        for page in reader.pages:
            text += page.extract_text() or "" # Tenta extrair texto da página
        return text.strip()
    except Exception as e:
        return f"ERRO ao extrair PDF: {e}"

def extract_text_from_docx(file_path):
    try:
        document = Document(file_path)
        return "\n".join([paragraph.text for paragraph in document.paragraphs])
    except Exception as e:
        return f"ERRO ao extrair DOCX: {e}"

def extract_text_from_xlsx(file_path):
    try:
        workbook = openpyxl.load_workbook(file_path)
        text = ""
        # Itera sobre todas as planilhas
        for sheet in workbook.sheetnames:
            worksheet = workbook[sheet]
            text += f"\n--- Planilha: {sheet} ---\n"
            # Lê as primeiras 10 linhas para demonstração
            for row in worksheet.iter_rows(max_row=10):
                text += " | ".join([str(cell.value) if cell.value is not None else "" for cell in row]) + "\n"
        return text.strip()
    except Exception as e:
        return f"ERRO ao extrair XLSX: {e}"

# Função que simula o envio para a fila e sistema OCR
def process_document_for_ocr(file_name, file_path):
    """
    Simula o fluxo de envio para a fila, OCR e salvamento no BD.
    Na implementação real, aqui você usaria ferramentas como RabbitMQ/Redis
    para a fila e um serviço/API de OCR (como Tesseract, Google Vision, etc.).
    """
    print(f"\n[Fase OCR/BD] Processando anexo: {file_name}")

    file_extension = os.path.splitext(file_name)[1].lower()
    
    # 1. Simula a chamada do OCR (aqui fazemos uma extração básica de texto)
    if file_extension == '.pdf':
        document_content = extract_text_from_pdf(file_path)
    elif file_extension == '.docx':
        document_content = extract_text_from_docx(file_path)
    elif file_extension == '.xlsx':
        document_content = extract_text_from_xlsx(file_path)
    else:
        document_content = "Tipo de arquivo não suportado para OCR/Processamento."

    print(f"   -> Conteúdo Capturado (Amostra):\n   {document_content[:200]}...")

    # 2. Simula o envio ao Banco de Dados PostgreSQL
    # Na implementação final, o resultado do OCR seria estruturado e salvo.
    print("   -> Informações processadas e PRONTAS para envio ao PostgreSQL.")
    
    # Exemplo (se você tivesse um modelo Django chamado Documento):
    # from automacao.models import Documento
    # Documento.objects.create(
    #     nome_arquivo=file_name,
    #     conteudo_extraido=document_content,
    #     status='Processado'
    # )
    
    print(f"[SUCESSO] Documento {file_name} finalizou o ciclo de automação.")